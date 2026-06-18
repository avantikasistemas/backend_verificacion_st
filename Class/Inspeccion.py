from Utils.tools import Tools, CustomException
from Utils.querys import Querys
from fastapi import Response
import pandas as pd
from io import BytesIO

class Inspeccion:

    def __init__(self, db):
        self.db = db
        self.tools = Tools()
        self.querys = Querys(self.db)

    # Función para cargar los datos registrados
    def obtener_tipo_inspeccion(self):
        """ Api que retorna los tipos de inspección activos. """
        try:
            data = self.querys.obtener_tipo_inspeccion()

            return self.tools.output(200, "Datos encontrados.", data)

        except CustomException as e:
            print(f"Error al obtener tipos de inspección: {e}")
            raise CustomException(f"{e}")

    # Función para obtener las modalidades de inspección
    def obtener_modalidad_inspeccion(self):
        """ Api que retorna las modalidades de inspección activas (Aéreo, Marítimo). """
        try:
            data = self.querys.obtener_modalidad_inspeccion()

            return self.tools.output(200, "Datos encontrados.", data)

        except CustomException as e:
            print(f"Error al obtener modalidades de inspección: {e}")
            raise CustomException(f"{e}")

    # Función para obtener aspectos según tipo de inspección
    def obtener_aspectos_por_tipo_inspeccion(self, data: dict):
        """ Api que retorna los aspectos de carga según el tipo de inspección. """
        try:
            tipo_inspeccion_id = data.get("tipo_inspeccion_id")
            data = self.querys.obtener_aspectos_por_tipo_inspeccion(tipo_inspeccion_id)
            
            return self.tools.output(200, "Aspectos obtenidos correctamente.", data)

        except CustomException as e:
            print(f"Error al obtener aspectos: {e}")
            raise CustomException(f"{e}")

    # Función para guardar inspección de carga
    def guardar_carga(self, data: dict):
        """ Api que guarda una inspección de carga con aspectos dinámicos e imágenes. """
        try:
            resultado = self.querys.guardar_carga(data)
            
            if resultado:
                return self.tools.output(200, "Inspección guardada correctamente.", None)
            else:
                return self.tools.output(500, "Error al guardar la inspección.", None)

        except CustomException as e:
            print(f"Error al guardar inspección: {e}")
            raise CustomException(f"{e}")

    # Función para cargar datos de inspecciones guardadas
    def cargar_datos_carga(self, data: dict):
        """ Api que retorna las inspecciones guardadas con sus aspectos e imágenes. """
        try:
            # Si flag_excel es True, exportar a Excel
            if data.get("flag_excel", False):
                resultado = self.querys.cargar_datos_carga(data)
                excel_data = self.exportar_excel_carga(resultado["registros"], data.get("tipo_inspeccion_id"))
                return Response(
                    content=excel_data["output"].read(),
                    headers=excel_data["headers"],
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            resultado = self.querys.cargar_datos_carga(data)
            
            return self.tools.output(200, "Datos cargados correctamente.", resultado)

        except CustomException as e:
            print(f"Error al cargar datos de inspección: {e}")
            raise CustomException(f"{e}")

    # Función para exportar datos de inspecciones de carga a Excel
    def exportar_excel_carga(self, registros: list, tipo_inspeccion_id: int):
        """
        Exporta los registros de inspecciones de carga a un archivo Excel
        con formato similar a la imagen proporcionada.
        
        Args:
            registros: Lista de inspecciones con sus aspectos
            tipo_inspeccion_id: ID del tipo de inspección para filtrar
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import CellIsRule

            # Convierte a mayúsculas cualquier valor de texto (deja intactos None/números)
            def mayus(valor):
                return valor.upper() if isinstance(valor, str) else valor
            
            # Crear un nuevo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Inspecciones"
            
            # Definir estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
            subheader_font = Font(bold=True, size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Obtener el tipo de inspección para el título
            tipo_nombre = registros[0]["tipo_inspeccion_nombre"] if registros else "Inspección"
            
            # Título principal
            ws.merge_cells('A1:Z1')
            ws['A1'] = f"LISTA DE CHEQUEO - {tipo_nombre.upper()}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = center_alignment
            ws['A1'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Si no hay registros, retornar archivo vacío
            if not registros:
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                headers = {
                    "Content-Disposition": f"attachment; filename=inspeccion_{tipo_nombre.replace(' ', '_')}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                }
                return {"output": output, "headers": headers}
            
            # Obtener todos los aspectos únicos y sus secciones del primer registro
            aspectos_header = []
            for seccion in registros[0].get("aspectos", []):
                for aspecto in seccion["aspectos"]:
                    aspectos_header.append({
                        "seccion": seccion["seccion_nombre"],
                        "aspecto": aspecto["aspecto_nombre"]
                    })
            
            # Fila 3: Encabezados de columnas principales
            current_row = 3
            col_index = 1
            
            # Columnas fijas iniciales
            columnas_fijas = []
            
            # Agregar columnas según el tipo de inspección (con FECHA primero)
            if tipo_inspeccion_id == 1:  # Carga Suelta
                columnas_fijas = [
                    "FECHA",
                    "REGISTRO",
                    "ADUANA",
                    "RESPONSABLE",
                    "USUARIO",
                    "NUMERO CONTENEDOR",
                    "N° SELLO SEGURIDAD",
                    "DOCUMENTO TRANSPORTE"
                ]
            elif tipo_inspeccion_id == 2:  # Contenedores
                columnas_fijas = [
                    "FECHA",
                    "REGISTRO",
                    "ADUANA",
                    "RESPONSABLE",
                    "USUARIO",
                    "EMPRESA TRANSPORTE",
                    "NUMERO CONTENEDOR",
                    "PLACA VEHICULO",
                    "PLACA TRAILER"
                ]
            else:
                columnas_fijas = [
                    "FECHA",
                    "REGISTRO",
                    "ADUANA",
                    "RESPONSABLE",
                    "USUARIO"
                ]
            
            # Escribir columnas fijas
            for col_name in columnas_fijas:
                cell = ws.cell(row=current_row, column=col_index)
                cell.value = col_name
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
                col_index += 1
            
            # Agrupar aspectos por sección para los encabezados
            secciones_agrupadas = {}
            for item in aspectos_header:
                seccion = item["seccion"]
                if seccion not in secciones_agrupadas:
                    secciones_agrupadas[seccion] = []
                secciones_agrupadas[seccion].append(item["aspecto"])
            
            # Fila 2: Encabezados de sección (merge cells)
            current_col = len(columnas_fijas) + 1
            for seccion, aspectos in secciones_agrupadas.items():
                start_col = current_col
                end_col = current_col + len(aspectos) - 1
                
                # Merge cells para el nombre de la sección
                if start_col == end_col:
                    cell = ws.cell(row=2, column=start_col)
                else:
                    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
                    cell = ws.cell(row=2, column=start_col)
                
                cell.value = seccion.upper()
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.alignment = center_alignment
                cell.border = border
                
                current_col += len(aspectos)
            
            # Fila 3: Nombres de aspectos individuales con numeración jerárquica
            seccion_counter = {}
            for item in aspectos_header:
                cell = ws.cell(row=current_row, column=col_index)
                
                # Obtener el índice de la sección
                seccion = item["seccion"]
                if seccion not in seccion_counter:
                    seccion_counter[seccion] = {
                        "numero_seccion": len(seccion_counter) + 1,
                        "contador_aspecto": 0
                    }
                
                seccion_counter[seccion]["contador_aspecto"] += 1
                
                # Crear numeración jerárquica (ej: 1.1, 1.2, 2.1, 2.2, etc.)
                numero_seccion = seccion_counter[seccion]["numero_seccion"]
                numero_aspecto = seccion_counter[seccion]["contador_aspecto"]
                cell.value = f"{numero_seccion}.{numero_aspecto}"
                
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
                col_index += 1

            # Rango de columnas de aspectos (para la fórmula de ESTADO DE LA INSPECCIÓN)
            col_aspecto_inicio = len(columnas_fijas) + 1
            col_aspecto_fin = col_index - 1
            letra_aspecto_inicio = get_column_letter(col_aspecto_inicio)
            letra_aspecto_fin = get_column_letter(col_aspecto_fin)

            # Columnas finales (FECHA ya está en columnas_fijas)
            columnas_finales = ["NOVEDADES", "ESTADO DE LA INSPECCIÓN"]
            estado_col_index = col_index + 1
            for col_name in columnas_finales:
                cell = ws.cell(row=current_row, column=col_index)
                cell.value = col_name
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
                col_index += 1
            
            # Escribir datos de cada registro
            current_row = 4
            for registro in registros:
                col_index = 1
                
                # Escribir columnas fijas (con FECHA primero)
                if tipo_inspeccion_id == 1:  # Carga Suelta
                    valores_fijos = [
                        registro.get("fecha_creacion", ""),
                        registro["id"],
                        registro.get("aduana_nombre", "N/A"),
                        registro.get("usuario_nombre") or registro.get("usuario", "N/A"),
                        registro.get("usuario", "N/A"),
                        registro.get("numero_contenedor", "N/A"),
                        registro.get("numero_sello_seguridad", "N/A"),
                        registro.get("documento_transporte", "N/A")
                    ]
                elif tipo_inspeccion_id == 2:  # Contenedores
                    valores_fijos = [
                        registro.get("fecha_creacion", ""),
                        registro["id"],
                        registro.get("aduana_nombre", "N/A"),
                        registro.get("usuario_nombre") or registro.get("usuario", "N/A"),
                        registro.get("usuario", "N/A"),
                        registro.get("empresa_transporte", "N/A"),
                        registro.get("numero_contenedor", "N/A"),
                        registro.get("placa_vehiculo", "N/A"),
                        registro.get("placa_trailer", "N/A")
                    ]
                else:
                    valores_fijos = [
                        registro.get("fecha_creacion", ""),
                        registro["id"],
                        registro.get("aduana_nombre", "N/A"),
                        registro.get("usuario_nombre") or registro.get("usuario", "N/A"),
                        registro.get("usuario", "N/A")
                    ]
                
                for valor in valores_fijos:
                    cell = ws.cell(row=current_row, column=col_index)
                    cell.value = mayus(valor) if valor else "N/A"
                    cell.alignment = center_alignment
                    cell.border = border
                    col_index += 1
                
                # Crear diccionario de aspectos del registro actual para búsqueda rápida
                aspectos_dict = {}
                for seccion in registro.get("aspectos", []):
                    for aspecto in seccion["aspectos"]:
                        aspectos_dict[aspecto["aspecto_nombre"]] = aspecto["valor"]
                
                # Escribir valores de aspectos en el orden correcto
                for item in aspectos_header:
                    cell = ws.cell(row=current_row, column=col_index)
                    valor = aspectos_dict.get(item["aspecto"], "N/A")
                    cell.value = valor
                    cell.alignment = center_alignment
                    cell.border = border
                    
                    # Aplicar color según el valor
                    if valor == "SI":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif valor == "NO":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    col_index += 1

                # Escribir novedades (FECHA ya no va al final)
                cell = ws.cell(row=current_row, column=col_index)
                cell.value = mayus(registro.get("novedades", ""))
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border
                col_index += 1

                # Estado de la inspección: fórmula que revisa si hubo algún "NO" en los aspectos
                cell = ws.cell(row=current_row, column=col_index)
                cell.value = (
                    f'=IF(COUNTIF({letra_aspecto_inicio}{current_row}:{letra_aspecto_fin}{current_row},"NO")>0,'
                    f'"NO FAVORABLE","FAVORABLE")'
                )
                cell.font = Font(bold=True)
                cell.alignment = center_alignment
                cell.border = border

                current_row += 1

            # Resaltar visualmente el resultado de "ESTADO DE LA INSPECCIÓN"
            estado_col_letter = get_column_letter(estado_col_index)
            if current_row > 4:
                rango_estado = f"{estado_col_letter}4:{estado_col_letter}{current_row - 1}"
                ws.conditional_formatting.add(
                    rango_estado,
                    CellIsRule(operator='equal', formula=['"FAVORABLE"'],
                               fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
                )
                ws.conditional_formatting.add(
                    rango_estado,
                    CellIsRule(operator='equal', formula=['"NO FAVORABLE"'],
                               fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
                )
            
            # Ajustar ancho de columnas
            for col in range(1, col_index + 1):
                column_letter = get_column_letter(col)
                if col <= len(columnas_fijas):
                    ws.column_dimensions[column_letter].width = 20
                elif col > col_index - 2:
                    ws.column_dimensions[column_letter].width = 25
                else:
                    ws.column_dimensions[column_letter].width = 8
            
            # Ajustar altura de filas
            ws.row_dimensions[1].height = 25
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 30
            
            # Guardar en BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            headers = {
                "Content-Disposition": f"attachment; filename=inspeccion_{tipo_nombre.replace(' ', '_')}.xlsx",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
            
            return {"output": output, "headers": headers}
            
        except Exception as e:
            print(f"Error al exportar Excel: {str(e)}")
            raise CustomException(f"Error al exportar Excel: {str(e)}")

    # Función para generar el PDF del detalle de una inspección de carga
    def generar_pdf_inspeccion_carga(self, data: dict):
        """ Genera un PDF con el detalle de una inspección de carga (igual al modal de detalle). """
        try:
            inspeccion_id = data.get("id")
            detalle = self.querys.obtener_detalle_inspeccion_carga(inspeccion_id)

            pdf_bytes = self._construir_pdf_detalle_carga(detalle)

            headers = {
                "Content-Disposition": f"attachment; filename=inspeccion_{inspeccion_id}.pdf",
            }
            return Response(
                content=pdf_bytes,
                headers=headers,
                media_type="application/pdf"
            )

        except CustomException as e:
            print(f"Error al generar PDF de inspección: {e}")
            raise CustomException(f"{e}")

    # Función auxiliar que construye el PDF del detalle con reportlab
    def _construir_pdf_detalle_carga(self, detalle: dict):
        from pathlib import Path
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage

        output = BytesIO()
        doc = SimpleDocTemplate(
            output, pagesize=letter,
            topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm
        )
        ancho_disponible = doc.width

        styles = getSampleStyleSheet()
        titulo_style = ParagraphStyle('Titulo', parent=styles['Heading1'], fontSize=14, spaceAfter=12)
        subtitulo_style = ParagraphStyle('Subtitulo', parent=styles['Heading2'], fontSize=11, spaceAfter=6)
        normal_style = styles['Normal']
        label_style = ParagraphStyle('Label', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=12)
        valor_style = ParagraphStyle('Valor', parent=styles['Normal'], fontSize=9, leading=12)
        aspecto_style = ParagraphStyle('Aspecto', parent=styles['Normal'], fontSize=8.5, leading=11)
        resultado_style = ParagraphStyle('Resultado', parent=styles['Normal'], fontSize=9, leading=11, alignment=1)

        # Convierte a mayúsculas cualquier valor de texto (deja intactos None/números)
        def mayus(valor):
            return valor.upper() if isinstance(valor, str) else valor

        elementos = []

        # Encabezado con el logotipo a la izquierda y el título al lado
        logo_path = Path.cwd() / "logotipo.png"
        if logo_path.exists():
            lector_logo = ImageReader(str(logo_path))
            ancho_logo_original, alto_logo_original = lector_logo.getSize()
            ancho_logo = 3.5 * cm
            alto_logo = ancho_logo * (alto_logo_original / ancho_logo_original)
            tabla_encabezado = Table(
                [[RLImage(str(logo_path), width=ancho_logo, height=alto_logo),
                  Paragraph(f"Detalle de Inspección #{detalle.get('id')}", titulo_style)]],
                colWidths=[ancho_logo + 0.3*cm, ancho_disponible - ancho_logo - 0.3*cm]
            )
            tabla_encabezado.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (0, 0), 0),
            ]))
            elementos.append(tabla_encabezado)
            elementos.append(Spacer(1, 10))
        else:
            elementos.append(Paragraph(f"Detalle de Inspección #{detalle.get('id')}", titulo_style))

        # Información general (layout de 2 columnas: etiqueta / valor, una pareja por fila)
        elementos.append(Paragraph("Información General", subtitulo_style))
        pares_info = [
            ("Tipo de Inspección:", mayus(detalle.get("tipo_inspeccion_nombre")) or "N/A"),
            ("Fecha de Creación:", detalle.get("fecha_creacion") or "N/A"),
        ]
        if detalle.get("modalidad_nombre"):
            pares_info.append(("Modalidad:", mayus(detalle.get("modalidad_nombre"))))
            pares_info.append(("Fecha y Hora Inicio:", detalle.get("fecha_hora_inicio") or "N/A"))
            pares_info.append(("Fecha y Hora Final:", detalle.get("fecha_hora_final") or "N/A"))

        pares_info.append(("Persona que realizó la inspección:", mayus(detalle.get("usuario_nombre")) or "N/A"))

        es_modalidad_aerea = (detalle.get("modalidad_nombre") or "").strip().lower() == "aéreo"

        if detalle.get("tipo_inspeccion_id") == 1:
            if not es_modalidad_aerea:
                pares_info.append(("Número de Contenedor:", mayus(detalle.get("numero_contenedor")) or "N/A"))
                pares_info.append(("N° Sello de Seguridad:", mayus(detalle.get("numero_sello_seguridad")) or "N/A"))
            pares_info.append(("Documento de Transporte:", mayus(detalle.get("documento_transporte")) or "N/A"))
        elif detalle.get("tipo_inspeccion_id") == 2:
            pares_info.append(("Empresa de Transporte Terrestre:", mayus(detalle.get("empresa_transporte")) or "N/A"))
            pares_info.append(("Número de Contenedor:", mayus(detalle.get("numero_contenedor")) or "N/A"))
            pares_info.append(("Placa de Vehículo:", mayus(detalle.get("placa_vehiculo")) or "N/A"))
            pares_info.append(("Placa de Trailer:", mayus(detalle.get("placa_trailer")) or "N/A"))

        if detalle.get("aduana_nombre"):
            pares_info.append(("Aduana:", mayus(detalle.get("aduana_nombre"))))

        # Se acomodan en una grilla de 2 parejas (4 columnas) por fila para no desperdiciar espacio,
        # pero cada celda envuelve su propio texto en un Paragraph para que haga wrap si no cabe.
        filas_info = []
        for i in range(0, len(pares_info), 2):
            fila = []
            for label, valor in pares_info[i:i+2]:
                fila.append(Paragraph(label, label_style))
                fila.append(Paragraph(str(valor), valor_style))
            if len(fila) == 2:
                fila.extend(["", ""])
            filas_info.append(fila)

        col_etiqueta = ancho_disponible * 0.18
        col_valor = ancho_disponible * 0.32
        tabla_info = Table(filas_info, colWidths=[col_etiqueta, col_valor, col_etiqueta, col_valor])
        tabla_info.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
        ]))
        elementos.append(tabla_info)
        elementos.append(Spacer(1, 10))

        elementos.append(Paragraph("Novedades", subtitulo_style))
        elementos.append(Paragraph(mayus(detalle.get("novedades")) or "SIN NOVEDADES", normal_style))
        elementos.append(Spacer(1, 14))

        # Aspectos evaluados
        elementos.append(Paragraph("Aspectos Evaluados", subtitulo_style))
        col_aspecto = ancho_disponible * 0.80
        col_resultado = ancho_disponible * 0.20
        for idx_seccion, seccion in enumerate(detalle.get("aspectos", []), start=1):
            elementos.append(Paragraph(f"{idx_seccion}. {seccion['seccion_nombre']}", styles['Heading3']))
            filas = [[
                Paragraph("Aspecto", ParagraphStyle('AspectoHeader', parent=aspecto_style, textColor=colors.white, fontName='Helvetica-Bold')),
                Paragraph("Resultado", ParagraphStyle('ResultadoHeader', parent=resultado_style, textColor=colors.white, fontName='Helvetica-Bold')),
            ]]
            for idx_aspecto, aspecto in enumerate(seccion["aspectos"], start=1):
                color_resultado = colors.black
                if aspecto['valor'] == 'SI':
                    color_resultado = colors.HexColor('#1e8449')
                elif aspecto['valor'] == 'NO':
                    color_resultado = colors.HexColor('#c0392b')
                filas.append([
                    Paragraph(f"{idx_seccion}.{idx_aspecto} {aspecto['aspecto_nombre']}", aspecto_style),
                    Paragraph(aspecto['valor'], ParagraphStyle('ResultadoValor', parent=resultado_style, textColor=color_resultado, fontName='Helvetica-Bold')),
                ])
            tabla_aspectos = Table(filas, colWidths=[col_aspecto, col_resultado], repeatRows=1)
            tabla_aspectos.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#aed6f1')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            elementos.append(tabla_aspectos)
            elementos.append(Spacer(1, 10))

        # Imágenes adjuntas
        imagenes = detalle.get("imagenes") or []
        if imagenes:
            elementos.append(Spacer(1, 6))
            elementos.append(Paragraph(f"Imágenes Adjuntas ({len(imagenes)})", subtitulo_style))
            uploads_dir = Path.cwd() / "Uploads"
            ancho_imagen = ancho_disponible * 0.6
            for img in imagenes:
                ruta_completa = uploads_dir / img.get("ruta_archivo", "")
                try:
                    if not ruta_completa.exists():
                        raise FileNotFoundError(str(ruta_completa))
                    lector = ImageReader(str(ruta_completa))
                    ancho_original, alto_original = lector.getSize()
                    alto_imagen = ancho_imagen * (alto_original / ancho_original)
                    elementos.append(RLImage(str(ruta_completa), width=ancho_imagen, height=alto_imagen))
                    elementos.append(Spacer(1, 10))
                except Exception as img_ex:
                    print(f"No se pudo incluir la imagen {img.get('ruta_archivo')} en el PDF: {img_ex}")
                    continue

        doc.build(elementos)
        output.seek(0)
        return output.read()

    # Función para obtener las aduanas activas
    def obtener_aduanas(self):
        """ Api que retorna las aduanas activas. """
        try:
            result = self.querys.obtener_aduanas()

            return self.tools.output(200, "Aduanas obtenidas correctamente.", result)

        except CustomException as e:
            print(f"Error al obtener aduanas: {e}")
            raise CustomException(f"{e}")

    # Función para obtener las aduanas según la modalidad
    def obtener_aduanas_por_modalidad(self, data: dict):
        """ Api que retorna las aduanas asociadas a una modalidad (Aéreo/Marítimo). """
        try:
            modalidad_id = data.get("modalidad_id")
            result = self.querys.obtener_aduanas_por_modalidad(modalidad_id)

            return self.tools.output(200, "Aduanas obtenidas correctamente.", result)

        except CustomException as e:
            print(f"Error al obtener aduanas por modalidad: {e}")
            raise CustomException(f"{e}")

    # Función para obtener responsables por aduana
    def obtener_responsables_por_aduana(self, data: dict):
        """ Api que retorna los responsables asociados a una aduana. """
        try:
            aduana_id = data.get("aduana_id")
            data = self.querys.obtener_responsables_por_aduana(aduana_id)
            
            return self.tools.output(200, "Responsables obtenidos correctamente.", data)

        except CustomException as e:
            print(f"Error al obtener responsables por aduana: {e}")
            raise CustomException(f"{e}")

