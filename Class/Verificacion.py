from Utils.tools import Tools, CustomException
from Utils.querys import Querys
from fastapi import Response
import pandas as pd
from io import BytesIO

class Verificacion:

    def __init__(self, db):
        self.db = db
        self.tools = Tools()
        self.querys = Querys(self.db)

    # Función guardar masivo
    def guardar_verificacion(self, data: dict):
        """ Api que realiza el guardado de los datos con aspectos dinámicos. """
        try:
            # Validar que los campos principales existan
            if not data.get("lugar_inspeccion_id"):
                raise CustomException("El campo lugar_inspeccion_id es requerido.")

            if not data.get("responsable_verificacion_id"):
                raise CustomException("El campo responsable_verificacion_id es requerido.")

            if not data.get("novedades"):
                raise CustomException("El campo novedades es requerido.")

            # Validar aspectos dinámicos
            aspectos_generales_dinamicos = data.get("aspectos_generales_dinamicos", {})
            if not aspectos_generales_dinamicos:
                raise CustomException("Debe completar al menos un aspecto de verificación.")
            
            # Validar que todos los aspectos tengan valores válidos
            for key, value in aspectos_generales_dinamicos.items():
                if value is None or value == "":
                    raise CustomException(
                        f"Todos los aspectos deben ser evaluados. El aspecto '{key}' no tiene valor."
                    )
                
                # Verificar que sea uno de los valores permitidos
                value_str = str(value)
                if value_str not in ["0", "1", "2"]:
                    raise CustomException(
                        f"El aspecto '{key}' tiene un valor inválido '{value}'. "
                        f"Solo se permiten los valores: 0 (NO), 1 (SI) o 2 (N/A)."
                    )
            
            # Guardamos la información en la base de datos.
            self.querys.guardar_verificacion(data)

            # Retornamos la información.
            return self.tools.output(200, "Datos guardados con éxito.", data)

        except CustomException as e:
            print(f"Error al guardar verificación: {e}")
            raise CustomException(f"{e}")
    
    # Función auxiliar para validar aspectos
    def _validar_aspectos(self, aspectos: dict, nombre_seccion: str, cantidad_esperada: int):
        """
        Valida que todos los aspectos de una sección tengan valores válidos (0, 1, o 2).
        
        Args:
            aspectos: Diccionario con los aspectos a validar
            nombre_seccion: Nombre de la sección para mensajes de error
            cantidad_esperada: Cantidad de aspectos que debe tener la sección
        """
        if not aspectos:
            raise CustomException(f"La sección '{nombre_seccion}' es requerida.")
        
        # Verificar que tenga la cantidad correcta de aspectos
        if len(aspectos) != cantidad_esperada:
            raise CustomException(
                f"La sección '{nombre_seccion}' debe tener {cantidad_esperada} aspectos, pero tiene {len(aspectos)}."
            )
        
        # Validar cada aspecto
        for key, value in aspectos.items():
            # Verificar que no sea None o vacío
            if value is None or value == "":
                raise CustomException(
                    f"El campo '{key}' en la sección '{nombre_seccion}' no puede estar vacío o nulo. "
                    f"Debe tener un valor: 0 (NO), 1 (SI) o 2 (N/A)."
                )
            
            # Convertir a string para validación uniforme
            value_str = str(value)
            
            # Verificar que sea uno de los valores permitidos
            if value_str not in ["0", "1", "2"]:
                raise CustomException(
                    f"El campo '{key}' en la sección '{nombre_seccion}' tiene un valor inválido '{value}'. "
                    f"Solo se permiten los valores: 0 (NO), 1 (SI) o 2 (N/A)."
                )

    # Función para cargar los datos registrados
    def cargar_datos(self, data: dict):
        """ Api que realiza la carga de los datos. """
        try:
            
            if data["flag_excel"]:
                registros = self.querys.cargar_datos(data)
                datos_excel = self.exportar_excel(registros, data.get("lugar_inspeccion_id"))
                return Response(
                    content=datos_excel["output"].read(), 
                    headers=datos_excel["headers"], 
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if data["position"] <= 0:
                message = "El campo posición no es válido"
                raise CustomException(message)
            
            # Cargamos la información en la base de datos.
            datos = self.querys.cargar_datos(data)
            
            registros = datos["registros"]
            cant_registros = datos["cant_registros"]
            
            if not registros:
                message = "No hay listado de reportes que mostrar."
                return self.tools.output(200, message, data={
                "total_registros": 0,
                "total_pag": 0,
                "posicion_pag": 1,
                "registros": []
            })
                
            if cant_registros%data["limit"] == 0:
                total_pag = cant_registros//data["limit"]
            else:
                total_pag = cant_registros//data["limit"] + 1
                
            if total_pag < int(data["position"]):
                message = "La posición excede el número total de registros."
                return self.tools.output(200, message, data={
                "total_registros": 0,
                "total_pag": 0,
                "posicion_pag": 0,
                "registros": []
            })

            registros_dict = {
                "total_registros": cant_registros,
                "total_pag": total_pag,
                "posicion_pag": data["position"],
                "registros": registros
            }

            # Retornamos la información.
            return self.tools.output(200, "Datos cargados con éxito.", registros_dict)

        except CustomException as e:
            print(f"Error al cargar datos: {e}")
            raise CustomException(f"{e}")

    # Función que realiza la operacion de exportar a Excel con formato profesional
    def exportar_excel(self, registros: list, lugar_inspeccion_id: int = None):
        """
        Exporta los registros de verificaciones a un archivo Excel
        con formato similar a la imagen proporcionada.
        
        Args:
            registros: Lista de verificaciones con sus aspectos
            lugar_inspeccion_id: ID del lugar de inspección para filtrar
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Crear un nuevo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Verificaciones"
            
            # Definir estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
            subheader_font = Font(bold=True, size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Obtener el lugar de inspección para el título
            lugar_nombre = registros[0]["lugar_inspeccion"] if registros else "Inspección"
            
            # Título principal
            ws.merge_cells('A1:Z1')
            ws['A1'] = f"LISTA DE CHEQUEO - CONTROL FÍSICO Y SEGURIDAD ({lugar_nombre.upper()})"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = center_alignment
            ws['A1'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Si no hay registros, retornar archivo vacío
            if not registros:
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                headers = {
                    "Content-Disposition": f"attachment; filename=verificacion_{lugar_nombre.replace(' ', '_')}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                }
                return {"output": output, "headers": headers}
            
            # Obtener todos los aspectos únicos y sus secciones del primer registro
            aspectos_header = []
            for seccion in registros[0].get("aspectos_agrupados", []):
                for aspecto in seccion["aspectos"]:
                    aspectos_header.append({
                        "seccion": seccion["nombre"],
                        "aspecto": aspecto["nombre"]
                    })
            
            # Fila 3: Encabezados de columnas principales
            current_row = 3
            col_index = 1
            
            # Columnas fijas iniciales - FECHA primero
            columnas_fijas = [
                "FECHA",
                "REGISTRO",
                "LUGAR INSPECCIÓN",
                "RESPONSABLE",
                "USUARIO"
            ]
            
            # Escribir columnas fijas
            for col_name in columnas_fijas:
                cell = ws.cell(row=current_row, column=col_index)
                cell.value = col_name
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
                col_index += 1
            
            # Agrupar aspectos por sección para los encabezados
            secciones_agrupadas = {}
            for item in aspectos_header:
                seccion = item["seccion"]
                if seccion not in secciones_agrupadas:
                    secciones_agrupadas[seccion] = []
                secciones_agrupadas[seccion].append(item["aspecto"])
            
            # Fila 2: Encabezados de sección (merge cells)
            current_col = len(columnas_fijas) + 1
            for seccion, aspectos in secciones_agrupadas.items():
                start_col = current_col
                end_col = current_col + len(aspectos) - 1
                
                # Merge cells para el nombre de la sección
                if start_col == end_col:
                    cell = ws.cell(row=2, column=start_col)
                else:
                    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
                    cell = ws.cell(row=2, column=start_col)
                
                cell.value = seccion.upper()
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.alignment = center_alignment
                cell.border = border
                
                current_col += len(aspectos)
            
            # Fila 3: Nombres de aspectos individuales con numeración jerárquica
            seccion_counter = {}
            for item in aspectos_header:
                cell = ws.cell(row=current_row, column=col_index)
                
                # Obtener el índice de la sección
                seccion = item["seccion"]
                if seccion not in seccion_counter:
                    seccion_counter[seccion] = {
                        "numero_seccion": len(seccion_counter) + 1,
                        "contador_aspecto": 0
                    }
                
                seccion_counter[seccion]["contador_aspecto"] += 1
                
                # Crear numeración jerárquica (ej: 1.1, 1.2, 2.1, 2.2, etc.)
                numero_seccion = seccion_counter[seccion]["numero_seccion"]
                numero_aspecto = seccion_counter[seccion]["contador_aspecto"]
                cell.value = f"{numero_seccion}.{numero_aspecto}"
                
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
                col_index += 1
            
            # Columnas finales (FECHA ya está en columnas_fijas)
            columnas_finales = ["NOVEDADES"]
            for col_name in columnas_finales:
                cell = ws.cell(row=current_row, column=col_index)
                cell.value = col_name
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
                col_index += 1
            
            # Escribir datos de cada registro
            current_row = 4
            for registro in registros:
                col_index = 1
                
                # Escribir columnas fijas (con FECHA primero)
                valores_fijos = [
                    registro.get("fecha_creacion", ""),
                    registro["id"],
                    registro.get("lugar_inspeccion", "N/A"),
                    registro.get("responsable_verificacion", "N/A"),
                    registro.get("usuario", "N/A")
                ]
                
                for valor in valores_fijos:
                    cell = ws.cell(row=current_row, column=col_index)
                    cell.value = valor if valor else "N/A"
                    cell.alignment = center_alignment
                    cell.border = border
                    col_index += 1
                
                # Crear diccionario de aspectos del registro actual para búsqueda rápida
                aspectos_dict = {}
                for seccion in registro.get("aspectos_agrupados", []):
                    for aspecto in seccion["aspectos"]:
                        aspectos_dict[aspecto["nombre"]] = aspecto["valor"]
                
                # Escribir valores de aspectos en el orden correcto
                for item in aspectos_header:
                    cell = ws.cell(row=current_row, column=col_index)
                    valor = aspectos_dict.get(item["aspecto"], "N/A")
                    
                    cell.value = valor
                    cell.alignment = center_alignment
                    cell.border = border
                    
                    # Aplicar color según el valor
                    if valor == "SI":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif valor == "NO":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    col_index += 1
                
                # Escribir novedades (FECHA ya no va al final)
                cell = ws.cell(row=current_row, column=col_index)
                cell.value = registro.get("novedades", "")
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border
                
                current_row += 1
            
            # Ajustar ancho de columnas
            for col in range(1, col_index + 1):
                column_letter = get_column_letter(col)
                if col <= len(columnas_fijas):
                    ws.column_dimensions[column_letter].width = 20
                elif col > col_index - 2:
                    ws.column_dimensions[column_letter].width = 25
                else:
                    ws.column_dimensions[column_letter].width = 8
            
            # Ajustar altura de filas
            ws.row_dimensions[1].height = 25
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 30
            
            # Guardar en BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            headers = {
                "Content-Disposition": f"attachment; filename=verificacion_{lugar_nombre.replace(' ', '_')}.xlsx",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
            
            return {"output": output, "headers": headers}
            
        except Exception as e:
            print(f"Error al exportar Excel: {str(e)}")
            raise CustomException(f"Error al exportar Excel: {str(e)}")

    # Función para obtener los lugares de inspección
    def obtener_lugares_inspeccion(self):
        """ Api que obtiene los lugares de inspección activos. """
        try:
            # Obtenemos los lugares de inspección desde la base de datos
            lugares = self.querys.obtener_lugares_inspeccion()
            
            if not lugares:
                message = "No hay lugares de inspección disponibles."
                return self.tools.output(200, message, data=[])
            
            # Retornamos la información
            return self.tools.output(200, "Lugares de inspección cargados con éxito.", lugares)

        except CustomException as e:
            print(f"Error al obtener lugares de inspección: {e}")
            raise CustomException(f"{e}")

    # Función para obtener los responsables por lugar de inspección
    def obtener_responsables_por_lugar(self, data: dict):
        """ Api que obtiene los responsables asociados a un lugar de inspección. """
        try:
            lugar_id = data.get("lugar_id")
            
            if not lugar_id:
                message = "El campo lugar_id es requerido."
                raise CustomException(message)
            
            # Obtenemos los responsables desde la base de datos
            responsables = self.querys.obtener_responsables_por_lugar(lugar_id)
            
            if not responsables:
                message = "No hay responsables asociados a este lugar de inspección."
                return self.tools.output(200, message, data=[])
            
            # Retornamos la información
            return self.tools.output(200, "Responsables cargados con éxito.", responsables)

        except CustomException as e:
            print(f"Error al obtener responsables: {e}")
            raise CustomException(f"{e}")

    # Función para generar el PDF con el detalle de una verificación (igual al modal de detalle)
    def generar_pdf_verificacion(self, data: dict):
        """ Genera un PDF con el detalle de una verificación de infraestructura (igual al modal de detalle). """
        try:
            verificacion_id = data.get("id")
            detalle = self.querys.obtener_detalle_verificacion(verificacion_id)

            pdf_bytes = self._construir_pdf_detalle_verificacion(detalle)

            headers = {
                "Content-Disposition": f"attachment; filename=verificacion_{verificacion_id}.pdf",
            }
            return Response(
                content=pdf_bytes,
                headers=headers,
                media_type="application/pdf"
            )

        except CustomException as e:
            print(f"Error al generar PDF de verificación: {e}")
            raise CustomException(f"{e}")

    # Función auxiliar que construye el PDF del detalle con reportlab
    def _construir_pdf_detalle_verificacion(self, detalle: dict):
        from pathlib import Path
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage

        output = BytesIO()
        doc = SimpleDocTemplate(
            output, pagesize=letter,
            topMargin=0.7*cm, bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm
        )
        ancho_disponible = doc.width

        styles = getSampleStyleSheet()
        titulo_style = ParagraphStyle('Titulo', parent=styles['Heading1'], fontSize=14, spaceAfter=12, alignment=0)
        subtitulo_style = ParagraphStyle('Subtitulo', parent=styles['Heading2'], fontSize=11, spaceAfter=6)
        normal_style = styles['Normal']
        label_style = ParagraphStyle('Label', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=12)
        valor_style = ParagraphStyle('Valor', parent=styles['Normal'], fontSize=9, leading=12)
        aspecto_style = ParagraphStyle('Aspecto', parent=styles['Normal'], fontSize=8.5, leading=11)
        resultado_style = ParagraphStyle('Resultado', parent=styles['Normal'], fontSize=9, leading=11, alignment=1)

        # Convierte a mayúsculas cualquier valor de texto (deja intactos None/números)
        def mayus(valor):
            return valor.upper() if isinstance(valor, str) else valor

        elementos = []

        # Encabezado con el logotipo a la izquierda y el título al lado
        logo_path = Path.cwd() / "logotipo.png"
        if logo_path.exists():
            lector_logo = ImageReader(str(logo_path))
            ancho_logo_original, alto_logo_original = lector_logo.getSize()
            ancho_logo = 7 * cm
            alto_logo = ancho_logo * (alto_logo_original / ancho_logo_original)
            tabla_encabezado = Table(
                [[RLImage(str(logo_path), width=ancho_logo, height=alto_logo),
                  Paragraph("Detalle de Verificación", titulo_style)]],
                colWidths=[ancho_logo + 0.2*cm, ancho_disponible - ancho_logo - 0.2*cm]
            )
            tabla_encabezado.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (0, 0), 0),
            ]))
            elementos.append(tabla_encabezado)
            elementos.append(Spacer(1, 10))
        else:
            elementos.append(Paragraph("Detalle de Verificación", titulo_style))

        # Información general (layout de 2 columnas: etiqueta / valor, una pareja por fila)
        elementos.append(Paragraph("Información General", subtitulo_style))
        pares_info = [
            ("Lugar de Inspección:", mayus(detalle.get("lugar_inspeccion")) or "N/A"),
            ("Responsable:", mayus(detalle.get("responsable_verificacion")) or "N/A"),
            ("Fecha de Creación:", detalle.get("fecha_creacion") or "N/A"),
            ("Persona que realizó la verificación:", mayus(detalle.get("usuario_nombre")) or "N/A"),
        ]

        filas_info = []
        for i in range(0, len(pares_info), 2):
            fila = []
            for label, valor in pares_info[i:i+2]:
                fila.append(Paragraph(label, label_style))
                fila.append(Paragraph(str(valor), valor_style))
            if len(fila) == 2:
                fila.extend(["", ""])
            filas_info.append(fila)

        col_etiqueta = ancho_disponible * 0.18
        col_valor = ancho_disponible * 0.32
        tabla_info = Table(filas_info, colWidths=[col_etiqueta, col_valor, col_etiqueta, col_valor])
        tabla_info.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
        ]))
        elementos.append(tabla_info)
        elementos.append(Spacer(1, 10))

        elementos.append(Paragraph("Novedades", subtitulo_style))
        elementos.append(Paragraph(mayus(detalle.get("novedades")) or "SIN NOVEDADES", normal_style))
        elementos.append(Spacer(1, 14))

        # Aspectos evaluados
        elementos.append(Paragraph("Aspectos Evaluados", subtitulo_style))
        col_aspecto = ancho_disponible * 0.80
        col_resultado = ancho_disponible * 0.20
        for idx_seccion, seccion in enumerate(detalle.get("aspectos_agrupados", []), start=1):
            elementos.append(Paragraph(f"{idx_seccion}. {seccion['nombre']}", styles['Heading3']))
            filas = [[
                Paragraph("Aspecto", ParagraphStyle('AspectoHeader', parent=aspecto_style, textColor=colors.white, fontName='Helvetica-Bold')),
                Paragraph("Resultado", ParagraphStyle('ResultadoHeader', parent=resultado_style, textColor=colors.white, fontName='Helvetica-Bold')),
            ]]
            for idx_aspecto, aspecto in enumerate(seccion["aspectos"], start=1):
                color_resultado = colors.black
                if aspecto['valor'] == 'SI':
                    color_resultado = colors.HexColor('#1e8449')
                elif aspecto['valor'] == 'NO':
                    color_resultado = colors.HexColor('#c0392b')
                filas.append([
                    Paragraph(f"{idx_seccion}.{idx_aspecto} {aspecto['nombre']}", aspecto_style),
                    Paragraph(aspecto['valor'], ParagraphStyle('ResultadoValor', parent=resultado_style, textColor=color_resultado, fontName='Helvetica-Bold')),
                ])
            tabla_aspectos = Table(filas, colWidths=[col_aspecto, col_resultado], repeatRows=1)
            tabla_aspectos.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#aed6f1')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            elementos.append(tabla_aspectos)
            elementos.append(Spacer(1, 10))

        # Imágenes adjuntas
        imagenes = detalle.get("imagenes") or []
        if imagenes:
            elementos.append(Spacer(1, 6))
            elementos.append(Paragraph(f"Imágenes Adjuntas ({len(imagenes)})", subtitulo_style))
            uploads_dir = Path.cwd() / "Uploads"
            ancho_imagen = ancho_disponible * 0.6
            for img in imagenes:
                ruta_completa = uploads_dir / img.get("ruta_archivo", "")
                try:
                    if not ruta_completa.exists():
                        raise FileNotFoundError(str(ruta_completa))
                    lector = ImageReader(str(ruta_completa))
                    ancho_original, alto_original = lector.getSize()
                    alto_imagen = ancho_imagen * (alto_original / ancho_original)
                    elementos.append(RLImage(str(ruta_completa), width=ancho_imagen, height=alto_imagen))
                    elementos.append(Spacer(1, 10))
                except Exception as img_ex:
                    print(f"No se pudo incluir la imagen {img.get('ruta_archivo')} en el PDF: {img_ex}")
                    continue

        doc.build(elementos)
        output.seek(0)
        return output.read()

    # Función para obtener los aspectos por lugar de inspección
    def obtener_aspectos_por_lugar(self, data: dict):
        """ Api que obtiene los aspectos de infraestructura asociados a un lugar de inspección. """
        try:
            lugar_id = data.get("lugar_id")
            
            if not lugar_id:
                message = "El campo lugar_id es requerido."
                raise CustomException(message)
            
            # Obtenemos los aspectos desde la base de datos
            aspectos = self.querys.obtener_aspectos_por_lugar_inspeccion(lugar_id)
            
            if not aspectos:
                message = "No hay aspectos asociados a este lugar de inspección."
                return self.tools.output(200, message, data=[])
            
            # Retornamos la información
            return self.tools.output(200, "Aspectos cargados con éxito.", aspectos)

        except CustomException as e:
            print(f"Error al obtener aspectos: {e}")
            raise CustomException(f"{e}")
